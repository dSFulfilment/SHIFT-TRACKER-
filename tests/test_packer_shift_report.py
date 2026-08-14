#!/usr/bin/env python3
"""Tests for Dandenong South packer shift report."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook

from packer_shift_report.cli import main
from packer_shift_report.compute import build_report
from packer_shift_report.load import load_boxes_packed, load_intra_hour
from packer_shift_report.validate import ColumnValidationError, validate_headers
from packer_shift_report.constants import BOXES_REQUIRED_COLUMNS


FIXTURES = ROOT / "fixtures" / "packer-shift"


class ValidateTests(unittest.TestCase):
    def test_missing_column_names_file(self):
        with self.assertRaises(ColumnValidationError) as ctx:
            validate_headers(
                "Boxes_Packed_by_Worker.xlsx",
                ["Report Date", "Shift", "Pnp Worker Name"],
                BOXES_REQUIRED_COLUMNS,
            )
        msg = str(ctx.exception)
        self.assertIn("Boxes_Packed_by_Worker.xlsx", msg)
        self.assertIn("Primary Sku", msg)


class FixtureReportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.boxes, cls.boxes_dropped = load_boxes_packed(
            FIXTURES / "Boxes_Packed_by_Worker.xlsx"
        )
        cls.intra, _ = load_intra_hour(FIXTURES / "Intra_Hour_Floor_Performance.xlsx")
        cls.report = build_report(cls.boxes, cls.boxes_dropped, cls.intra)

    def test_blank_rows_dropped(self):
        self.assertGreaterEqual(self.boxes_dropped["blank_worker_or_sku"], 2)

    def test_short_lines_still_scored(self):
        names = {r.worker_display for r in self.report.morning}
        self.assertIn("Eve Short", names)
        by_name = {r.worker_display: r for r in self.report.morning}
        self.assertEqual(by_name["Eve Short"].flag, "Below target")
        self.assertFalse(hasattr(self.report.exclusions, "under_15_min") and self.report.exclusions.under_15_min)

    def test_morning_flags(self):
        by_name = {r.worker_display: r for r in self.report.morning}
        self.assertIn("Bob Jones", by_name)
        self.assertEqual(by_name["Bob Jones"].flag, "Below target")
        self.assertIn("short by", by_name["Bob Jones"].why.lower())
        self.assertTrue(by_name["Bob Jones"].sku_lines)
        self.assertIsNotNone(self.report.morning_totals)
        self.assertGreaterEqual(self.report.morning_totals.packers, 3)
        self.assertIn("Alice Smith", by_name)
        self.assertEqual(by_name["Alice Smith"].flag, "On/above target")
        self.assertAlmostEqual(by_name["Alice Smith"].pct_of_target, 130 / 110 * 100, places=5)
        self.assertIn("Carol Lee", by_name)
        self.assertEqual(by_name["Carol Lee"].flag, "On/above target")

    def test_afternoon_separate_and_unknown_sku(self):
        by_name = {r.worker_display: r for r in self.report.afternoon}
        self.assertIn("Alice Smith", by_name)  # not merged with morning
        self.assertIn("Dan West", by_name)
        self.assertEqual(by_name["Dan West"].flag, "No target defined")
        self.assertIn("999", by_name["Dan West"].notes)

    def test_sorted_worst_to_best(self):
        pcts = [r.pct_of_target for r in self.report.morning if r.pct_of_target is not None]
        self.assertEqual(pcts, sorted(pcts))

    def test_cli_writes_workbook(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            code = main(
                [
                    "--boxes",
                    str(FIXTURES / "Boxes_Packed_by_Worker.xlsx"),
                    "--intra",
                    str(FIXTURES / "Intra_Hour_Floor_Performance.xlsx"),
                    "--out",
                    str(out),
                ]
            )
            self.assertEqual(code, 0)
            self.assertTrue(out.is_file())
            wb = load_workbook(out)
            for title in (
                "How this works",
                "SKU targets",
                "Raw data",
                "Morning shift",
                "Afternoon shift",
                "Why (SKU detail)",
                "Exclusions",
                "Intra hour (reference)",
            ):
                self.assertIn(title, wb.sheetnames)
            raw = wb["Raw data"]
            self.assertTrue(str(raw["I2"].value).startswith("=IF("))
            morning = wb["Morning shift"]
            # Shift amount block at top
            self.assertEqual(morning["A1"].value, "Shift amount")
            # Packer table uses SUMIFS somewhere in column B
            found_sumifs = False
            for row in morning.iter_rows(min_col=2, max_col=2, min_row=1, max_row=40):
                val = row[0].value
                if isinstance(val, str) and val.startswith("=SUMIFS("):
                    found_sumifs = True
                    break
            self.assertTrue(found_sumifs, "Morning sheet should SUMIFS into Raw data")
            why = wb["Why (SKU detail)"]
            self.assertEqual(why["A1"].value, "Shift")
            self.assertIn("Line verdict", [c.value for c in why[1]])
            sku = wb["SKU targets"]
            self.assertEqual(sku["A2"].value, 125)
            self.assertEqual(sku["B2"].value, 17)

    def test_cli_dir_mode(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "dir-out.xlsx"
            code = main(["--dir", str(FIXTURES), "--out", str(out)])
            self.assertEqual(code, 0)
            self.assertTrue(out.is_file())

    def test_missing_column_fails_clearly(self):
        with tempfile.TemporaryDirectory() as td:
            bad = Path(td) / "Boxes_Packed_by_Worker.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Report Date", "Shift", "Pnp Worker Name", "Boxes Packed"])
            ws.append(["2026-08-13", "morning_shift", "Ada", 10])
            wb.save(bad)
            intra = Path(td) / "Intra_Hour_Floor_Performance.xlsx"
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.append(["Report Date Hour", "Pnp Worker Name", "Boxes Packed"])
            ws2.append(["2026-08-13 08:00", "Ada", 10])
            wb2.save(intra)
            code = main(
                [
                    "--boxes",
                    str(bad),
                    "--intra",
                    str(intra),
                    "--out",
                    str(Path(td) / "x.xlsx"),
                ]
            )
            self.assertEqual(code, 1)


class DipStrikeTests(unittest.TestCase):
    def test_dipped_below_strike_flag(self):
        # Overall on target, but one SKU under strike
        boxes = [
            {
                "Report Date": "2026-08-13",
                "Shift": "morning_shift",
                "Pnp Worker Name": "Pat Dip",
                "Station Name": "A1",
                "Primary Sku": 250,
                "Boxes Packed": 14,  # 1h → BPH 14 < strike 14.6
                "Packing Time Seconds": 3600,
            },
            {
                "Report Date": "2026-08-13",
                "Shift": "morning_shift",
                "Pnp Worker Name": "Pat Dip",
                "Station Name": "A1",
                "Primary Sku": 500,
                "Boxes Packed": 50,  # 1h → BPH 50 >> target 23
                "Packing Time Seconds": 3600,
            },
        ]
        report = build_report(boxes, {"blank_worker_or_sku": 0})
        # boxes 64, target boxes = 16*1 + 23*1 = 39, pct >> 100
        # but 250 line under strike → Dipped below strike
        self.assertEqual(len(report.morning), 1)
        self.assertEqual(report.morning[0].flag, "Dipped below strike")
        self.assertGreaterEqual(report.morning[0].pct_of_target, 100)


if __name__ == "__main__":
    unittest.main()
