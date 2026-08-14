"""Load the three shift exports. Never modifies source files."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .constants import (
    BOXES_REQUIRED_COLUMNS,
    INTRA_REQUIRED_COLUMNS,
    SUMMARY_REQUIRED_COLUMNS,
)
from .validate import canonicalize_header, find_header_row_index, validate_headers


def _sheet_rows(path: Path, required: List[str]) -> Tuple[List[str], List[List[Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not all_rows:
            return [], []
        header_idx = find_header_row_index(all_rows, required)
        headers = validate_headers(path.name, all_rows[header_idx], required)
        data = all_rows[header_idx + 1 :]
        return headers, data
    finally:
        wb.close()


def _row_dict(headers: List[str], row: List[Any]) -> Dict[str, Any]:
    out = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        out[h] = row[i] if i < len(row) else None
    return out


def _is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def normalize_worker_key(name: str) -> str:
    """Grouping key only — trim + collapse spaces. Display keeps original spelling."""
    return " ".join(str(name).strip().split())


def parse_sku(value) -> Optional[Any]:
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel may store 250.0
        if float(value) == int(value):
            return int(value)
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        f = float(s)
        if f == int(f):
            return int(f)
        return f
    except ValueError:
        return s


def load_boxes_packed(path: Path, file_label: str = "Boxes_Packed_by_Worker.xlsx"):
    headers, raw_rows = _sheet_rows(path, BOXES_REQUIRED_COLUMNS)
    kept = []
    dropped = {"blank_worker_or_sku": 0}
    for row in raw_rows:
        d = _row_dict(headers, row)
        # Trailing blank / total rows — drop when Primary Sku or Pnp Worker Name blank
        if _is_blank(d.get("Pnp Worker Name")) or _is_blank(d.get("Primary Sku")):
            # Ignore fully empty trailing rows silently in the blank counter only if all key fields empty
            if all(_is_blank(d.get(c)) for c in ("Pnp Worker Name", "Primary Sku", "Boxes Packed")):
                continue
            dropped["blank_worker_or_sku"] += 1
            continue
        kept.append(d)
    return kept, dropped


def load_intra_hour(path: Path, file_label: str = "Intra_Hour_Floor_Performance.xlsx"):
    """
    Kept for a later "which hour did they slow down" view.
    Not used in the core target comparison — missing file should not block if optional,
    but when provided we still validate columns.
    """
    headers, raw_rows = _sheet_rows(path, INTRA_REQUIRED_COLUMNS)
    kept = []
    dropped = {"blank_worker": 0}
    for row in raw_rows:
        d = _row_dict(headers, row)
        if _is_blank(d.get("Pnp Worker Name")):
            if all(_is_blank(d.get(c)) for c in ("Report Date Hour", "Pnp Worker Name", "Boxes Packed")):
                continue
            dropped["blank_worker"] += 1
            continue
        kept.append(d)
    return kept, dropped


def load_facility_summary(path: Path, file_label: str = "Overall_Summary_by_Packer_and_Date.xlsx"):
    """
    ONLY used to filter Facility Name == Dandenong South.

    Do NOT use its blended Boxes per Hour for target comparisons — it mixes every
    SKU a packer touched into one number, and targets vary a lot by SKU.
    """
    headers, raw_rows = _sheet_rows(path, SUMMARY_REQUIRED_COLUMNS)
    kept = []
    for row in raw_rows:
        d = _row_dict(headers, row)
        if _is_blank(d.get("Pnp Worker Name")):
            continue
        kept.append(d)
    return kept
