"""Build the auditable openpyxl workbook (formulas on Raw + shift sheets)."""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .compute import PackerShiftResult, ReportData
from .constants import MIN_HOURS_ON_SKU, SKU_TARGETS

# Sheet titles (used in formulas — keep stable)
SHEET_HOW = "How this works"
SHEET_SKU = "SKU targets"
SHEET_RAW = "Raw data"
SHEET_MORNING = "Morning shift"
SHEET_AFTERNOON = "Afternoon shift"
SHEET_WHY = "Why (SKU detail)"
SHEET_EXCLUSIONS = "Exclusions"
SHEET_INTRA = "Intra hour (reference)"

FILL_RED = PatternFill("solid", fgColor="FECACA")
FILL_AMBER = PatternFill("solid", fgColor="FDE68A")
FILL_GREEN = PatternFill("solid", fgColor="BBF7D0")
FILL_HEADER = PatternFill("solid", fgColor="1C2430")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_TITLE = Font(bold=True, size=14)
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header(ws, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autosize(ws, min_width=10, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _write_how(ws, data: ReportData):
    lines = [
        "Packer performance vs target — Dandenong South",
        "",
        "What this workbook tells you",
        "Each shift, packers work one or more SKUs. Every SKU has its own target boxes-per-hour (BPH)",
        "and a lower “strike” line. This report scores each packer for Morning and Afternoon separately,",
        "so you can see who hit target, who dipped under the strike line on at least one SKU, and who",
        "finished the shift under target overall.",
        "",
        "Where the numbers come from",
        "1. Boxes Packed by Worker — one row per packer + SKU + shift (this is the performance source).",
        "2. Intra Hour Floor Performance — stored on “Intra hour (reference)” for a later hour-by-hour",
        "   slowdown view. It does not affect the target comparison.",
        "Exports are assumed to be the Dandenong South shift files you selected.",
        "",
        "How each packer+SKU line is measured",
        f"• Hours on SKU = Packing Time Seconds ÷ 3600",
        "• Actual BPH = Boxes Packed ÷ Hours on SKU",
        f"• Lines under {MIN_HOURS_ON_SKU:.2f} hours (15 minutes) are excluded — that is usually changeover/",
        "  setup noise (1–2 slow boxes right after a SKU switch), not real shift performance.",
        "",
        "How the shift score is calculated (per packer, per shift)",
        "• % of target = (total Boxes Packed on included lines) ÷ (total of Hours×SKU Target BPH",
        "  on included lines that have a known target) × 100",
        "• Below target — % of target < 100",
        "• Dipped below strike — % of target ≥ 100, but at least one included line had Actual BPH",
        "  below that SKU’s strike line",
        "• On/above target — everything else",
        "A packer on both shifts appears twice (once on Morning, once on Afternoon) — shifts are never merged.",
        "",
        "Unknown SKUs",
        "If a Primary Sku is not on the SKU targets sheet, the line is kept and flagged",
        "(“no target defined for SKU X”). We do not invent a target.",
        "",
        "Auditability",
        "The Raw data sheet holds the export values plus Excel formulas for Hours, Actual BPH, Target BPH",
        "(INDEX/MATCH into SKU targets), Strike BPH, Target boxes, and the include flag.",
        "Morning / Afternoon sheets show the shift amount at the top, then each packer with a Why column.",
        "Why (SKU detail) breaks every packer into SKU lines so you can see what dragged or held the score.",
        "Morning / Afternoon metrics use SUMIFS against Raw data so totals stay auditable if you edit Raw.",
        "",
        "Source files are never modified — this workbook is a new output only.",
    ]
    for i, text in enumerate(lines, start=1):
        cell = ws.cell(i, 1, text)
        if i == 1:
            cell.font = FONT_TITLE
        elif text and not text.startswith("•") and text[0].isupper() and i > 1 and lines[i - 2] == "":
            cell.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 110
    if data.warnings:
        start = len(lines) + 2
        ws.cell(start, 1, "Warnings").font = Font(bold=True, color="B91C1C")
        for j, w in enumerate(data.warnings):
            ws.cell(start + 1 + j, 1, w)


def _write_sku(ws):
    ws.append(["Primary Sku", "Target BPH", "Strike line BPH"])
    for sku in sorted(SKU_TARGETS.keys()):
        target, strike = SKU_TARGETS[sku]
        ws.append([sku, target, strike])
    _style_header(ws, 3)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
        for c in row:
            c.border = THIN
    _autosize(ws)
    # Named-friendly range note in D1
    ws["E1"] = "Used by INDEX/MATCH on Raw data — edit here to change targets globally."
    ws["E1"].font = Font(italic=True, color="6B7280")


def _write_exclusions(ws, data: ReportData):
    ws.append(["Reason", "Row count"])
    for reason, count in data.exclusions.as_rows():
        ws.append([reason, count])
    ws.append([])
    ws.append(["Workers in Boxes export", len(data.facility_workers)])
    _style_header(ws, 2)
    _autosize(ws)


def _write_intra(ws, data: ReportData):
    ws.append(
        [
            "Report Date Hour",
            "Pnp Worker Name",
            "Boxes Packed",
            "Note",
        ]
    )
    if not data.intra_rows:
        ws.append(["", "", "", "No intra-hour rows loaded."])
    else:
        for r in data.intra_rows:
            ws.append(
                [
                    r.get("Report Date Hour"),
                    r.get("Pnp Worker Name"),
                    r.get("Boxes Packed"),
                    "Reference only — not used in % of target",
                ]
            )
    _style_header(ws, 4)
    _autosize(ws)


def _write_raw(ws, data: ReportData):
    headers = [
        "Report Date",
        "Shift key",
        "Shift",
        "Worker Name",
        "Station Name",
        "Primary Sku",
        "Boxes Packed",
        "Packing Time Seconds",
        "Hours on SKU",
        "Actual BPH",
        "Target BPH",
        "Strike BPH",
        "Target boxes",
        "Include flag",
        "Below strike flag",
        "Score boxes",
        "Exclude / note",
    ]
    ws.append(headers)
    sku_sheet = f"'{SHEET_SKU}'"
    for i, line in enumerate(data.raw_lines, start=2):
        # Values from source / loader
        ws.cell(i, 1, line.report_date)
        ws.cell(i, 2, line.shift_key)
        ws.cell(i, 3, line.shift_label)
        ws.cell(i, 4, line.worker_display)
        ws.cell(i, 5, line.station)
        ws.cell(i, 6, line.sku)
        ws.cell(i, 7, line.boxes if line.exclude_reason != "Missing Boxes Packed" else None)
        ws.cell(i, 8, line.packing_seconds if line.exclude_reason != "Missing Packing Time Seconds" else None)

        # Formulas — auditable (same rules as packer_shift_report.compute)
        ws.cell(i, 9, f'=IF(OR(H{i}="",H{i}=0),"",H{i}/3600)')
        ws.cell(i, 10, f'=IF(OR(I{i}="",I{i}=0),"",G{i}/I{i})')
        ws.cell(
            i,
            11,
            f'=IF(F{i}="","",IFERROR(INDEX({sku_sheet}!$B$2:$B$50,MATCH(F{i},{sku_sheet}!$A$2:$A$50,0)),"no target defined"))',
        )
        ws.cell(
            i,
            12,
            f'=IF(F{i}="","",IFERROR(INDEX({sku_sheet}!$C$2:$C$50,MATCH(F{i},{sku_sheet}!$A$2:$A$50,0)),"no target defined"))',
        )
        ws.cell(
            i,
            13,
            f'=IF(OR(I{i}="",K{i}="",K{i}="no target defined"),"",I{i}*K{i})',
        )
        # Include flag: 1 = long enough for scoring (unknown SKU still 1 — flagged in notes)
        ws.cell(i, 14, f'=IF(I{i}="","",IF(I{i}<{MIN_HOURS_ON_SKU},0,1))')
        ws.cell(
            i,
            15,
            f'=IF(OR(N{i}<>1,J{i}="",L{i}="",L{i}="no target defined"),0,IF(J{i}<L{i},1,0))',
        )
        # Score boxes: included lines with a known target only (drives % of target)
        ws.cell(
            i,
            16,
            f'=IF(OR(N{i}<>1,K{i}="",K{i}="no target defined"),0,G{i})',
        )
        note = line.exclude_reason
        if line.unknown_sku:
            note = (("" if not note else note + "; ") + f"no target defined for SKU {line.sku}").strip("; ")
        ws.cell(i, 17, note)

    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{max(1, ws.max_row)}"


def _flag_fill(flag: str):
    if flag == "Below target":
        return FILL_RED
    if flag == "Dipped below strike":
        return FILL_AMBER
    if flag == "On/above target":
        return FILL_GREEN
    return PatternFill("solid", fgColor="E5E7EB")


def _write_shift_sheet(ws, results: List[PackerShiftResult], shift_key: str, totals=None):
    # Shift amount summary — what the whole shift packed vs target
    if totals is not None:
        ws.append(["Shift amount", totals.shift_label])
        ws.append(["Packers scored", totals.packers])
        ws.append(["Hours worked (included)", round(totals.hours, 2)])
        ws.append(["Boxes packed (included)", round(totals.boxes, 1)])
        ws.append(["Target boxes (hours × SKU targets)", round(totals.target_boxes, 1)])
        ws.append(
            [
                "% of target (shift)",
                round(totals.pct_of_target, 1) if totals.pct_of_target is not None else "—",
            ]
        )
        gap = totals.box_gap
        ws.append(
            [
                "Box gap (ahead if +)",
                round(gap, 1) if gap is not None else "—",
            ]
        )
        ws.append(
            [
                "Flags",
                f"Below {totals.below} · Dip strike {totals.dipped} · On/above {totals.on_target} · No target {totals.no_target}",
            ]
        )
        ws.append([])
        ws["A1"].font = Font(bold=True, size=12)

    header_row = ws.max_row + 1
    headers = [
        "Packer",
        "Hours worked",
        "Boxes packed",
        "Target boxes",
        "% of target",
        "Box gap",
        "Flag",
        "Why it worked / didn't",
        "Notes",
    ]
    ws.append(headers)
    raw = f"'{SHEET_RAW}'"
    first_data = header_row + 1
    for offset, r in enumerate(results):
        i = first_data + offset
        ws.cell(i, 1, r.worker_display)
        ws.cell(
            i,
            2,
            f'=SUMIFS({raw}!$I:$I,{raw}!$D:$D,A{i},{raw}!$B:$B,"{shift_key}",{raw}!$N:$N,1)',
        )
        ws.cell(
            i,
            3,
            f'=SUMIFS({raw}!$G:$G,{raw}!$D:$D,A{i},{raw}!$B:$B,"{shift_key}",{raw}!$N:$N,1)',
        )
        ws.cell(
            i,
            4,
            f'=SUMIFS({raw}!$M:$M,{raw}!$D:$D,A{i},{raw}!$B:$B,"{shift_key}",{raw}!$N:$N,1)',
        )
        ws.cell(
            i,
            5,
            f'=IF(D{i}=0,"",SUMIFS({raw}!$P:$P,{raw}!$D:$D,A{i},{raw}!$B:$B,"{shift_key}",{raw}!$N:$N,1)/D{i}*100)',
        )
        ws.cell(i, 6, f'=IF(OR(D{i}="",E{i}=""),"",SUMIFS({raw}!$P:$P,{raw}!$D:$D,A{i},{raw}!$B:$B,"{shift_key}",{raw}!$N:$N,1)-D{i})')
        strike_sum = (
            f'SUMIFS({raw}!$O:$O,{raw}!$D:$D,A{i},{raw}!$B:$B,"{shift_key}",{raw}!$N:$N,1)'
        )
        ws.cell(
            i,
            7,
            f'=IF(D{i}=0,"No target defined",IF(E{i}<100,"Below target",IF({strike_sum}>0,"Dipped below strike","On/above target")))',
        )
        ws.cell(i, 8, r.why)
        ws.cell(i, 9, r.notes)

        fill = _flag_fill(r.flag)
        for c in range(1, 10):
            ws.cell(i, c).fill = fill
            ws.cell(i, c).border = THIN
        ws.cell(i, 2).number_format = "0.00"
        ws.cell(i, 3).number_format = "#,##0.00"
        ws.cell(i, 4).number_format = "#,##0.00"
        ws.cell(i, 5).number_format = "0.0"
        ws.cell(i, 6).number_format = "0.0"
        ws.cell(i, 8).alignment = Alignment(wrap_text=True, vertical="top")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, c, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
    legend_row = first_data + len(results) + 2
    ws.cell(legend_row, 1, "Colour key").font = Font(bold=True)
    ws.cell(legend_row + 1, 1, "Below target").fill = FILL_RED
    ws.cell(legend_row + 2, 1, "Dipped below strike").fill = FILL_AMBER
    ws.cell(legend_row + 3, 1, "On/above target").fill = FILL_GREEN
    ws.cell(
        legend_row + 5,
        1,
        "Click a packer name into the Why (SKU detail) sheet filter to see each SKU that dragged or held the score. "
        "Hours/boxes/target/% are SUMIFS into Raw data.",
    )
    ws.cell(legend_row + 5, 1).font = Font(italic=True, color="6B7280")
    ws.column_dimensions["H"].width = 56
    _autosize(ws)
    ws.freeze_panes = f"A{first_data}"


def _write_why_detail(ws, data: ReportData):
    ws.append(
        [
            "Shift",
            "Packer",
            "Flag",
            "Why (shift)",
            "SKU",
            "Hours",
            "Boxes",
            "Actual BPH",
            "Target BPH",
            "Strike BPH",
            "Target boxes",
            "Line % of target",
            "Line verdict",
        ]
    )
    for block in (data.morning, data.afternoon):
        for r in block:
            if not r.sku_lines:
                ws.append(
                    [
                        r.shift_label,
                        r.worker_display,
                        r.flag,
                        r.why,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )
                continue
            for line in r.sku_lines:
                ws.append(
                    [
                        r.shift_label,
                        r.worker_display,
                        r.flag,
                        r.why,
                        line.sku,
                        round(line.hours, 3),
                        line.boxes,
                        round(line.actual_bph, 2) if line.actual_bph is not None else None,
                        line.target_bph,
                        line.strike_bph,
                        round(line.target_boxes, 2) if line.target_boxes is not None else None,
                        round(line.line_pct, 1) if line.line_pct is not None else None,
                        line.verdict,
                    ]
                )
    _style_header(ws, 13)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=13):
        verdict = row[12].value
        fill = None
        if verdict == "under strike":
            fill = FILL_RED
        elif verdict == "under target":
            fill = FILL_AMBER
        elif verdict == "on/above target":
            fill = FILL_GREEN
        if fill:
            for c in row:
                c.fill = fill
        for c in row:
            c.border = THIN
    _autosize(ws)
    ws.auto_filter.ref = f"A1:M{max(1, ws.max_row)}"
    ws.freeze_panes = "A2"


def write_workbook(data: ReportData, out_path: Path) -> Path:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = SHEET_HOW
    _write_how(ws0, data)

    ws_sku = wb.create_sheet(SHEET_SKU)
    _write_sku(ws_sku)

    ws_raw = wb.create_sheet(SHEET_RAW)
    _write_raw(ws_raw, data)

    ws_m = wb.create_sheet(SHEET_MORNING)
    _write_shift_sheet(ws_m, data.morning, "morning_shift", data.morning_totals)

    ws_a = wb.create_sheet(SHEET_AFTERNOON)
    _write_shift_sheet(ws_a, data.afternoon, "afternoon_shift", data.afternoon_totals)

    ws_why = wb.create_sheet(SHEET_WHY)
    _write_why_detail(ws_why, data)

    ws_x = wb.create_sheet(SHEET_EXCLUSIONS)
    _write_exclusions(ws_x, data)

    ws_i = wb.create_sheet(SHEET_INTRA)
    _write_intra(ws_i, data)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
