#!/usr/bin/env python3
"""
strike_check.py — fair strike-candidate report from Boxes Packed + Raw Data.

Single-SKU and mixed-size data are calculated and shown separately — never blended.
Strike pass/fail comes only from Boxes_Packed_by_Worker (one SKU per row).
Raw_Data single-SKU sessions attach as condensed context; mixed (2+ sizes) is
rolled up one row per worker and never used to judge any one SKU.
Default screen view: strike candidates only (pass --all-rows / --mixed-all to expand).
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

# Optional deps — Boxes via openpyxl; Raw_Data via pandas (per spec).
try:
    import pandas as pd
except ImportError:  # pragma: no cover
    pd = None  # type: ignore

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore


# ---------------------------------------------------------------------------
# Reference strike table (override with --strike-table CSV)
# ---------------------------------------------------------------------------

DEFAULT_STRIKE_TABLE: Dict[int, Tuple[float, float]] = {
    # SKU: (Target BPH, Strike Line)
    125: (17.0, 15.7),
    150: (18.0, 16.3),
    200: (17.0, 15.3),
    250: (16.0, 14.6),
    300: (18.0, 16.0),
    400: (16.0, 14.6),
    500: (23.0, 20.6),
    600: (20.0, 17.8),
    700: (21.0, 19.0),
}

MIN_PACKING_SECONDS = 600  # 10 minutes — below → insufficient_data


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and v != v:  # NaN
        return True
    return str(v).strip() == ""


def _num(v: Any) -> Optional[float]:
    if _blank(v):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v) if float(v) == float(v) else None
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _norm_name(name: Any) -> str:
    """Trim + lowercase + collapse spaces for cross-file matching."""
    return " ".join(str(name).strip().lower().split())


def _parse_sku(v: Any) -> Optional[int]:
    if _blank(v):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        if f == int(f):
            return int(f)
        return None
    s = str(v).strip().lower().rstrip("g").strip()
    try:
        f = float(s)
        if f == int(f):
            return int(f)
    except ValueError:
        return None
    return None


def parse_box_sku_sizes(raw: Any) -> Tuple[List[int], str]:
    """
    Split Box Sku Sizes on ',', trim, strip trailing 'g'.
    Returns (sku_list, kind) where kind is 'single_sku' | 'mixed_sku' | 'empty'.
    """
    if _blank(raw):
        return [], "empty"
    s = str(raw).strip()
    # Drop surrounding quotes if present
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        s = s[1:-1]
    parts = [p.strip() for p in s.split(",") if p.strip()]
    skus: List[int] = []
    for p in parts:
        sku = _parse_sku(p)
        if sku is not None and sku not in skus:
            skus.append(sku)
    if len(skus) == 0:
        return [], "empty"
    if len(skus) == 1:
        return skus, "single_sku"
    return skus, "mixed_sku"


def load_strike_table(path: Optional[Path] = None) -> Dict[int, Tuple[float, float]]:
    """Hardcoded defaults, or CSV override: SKU, Target BPH, Strike Line."""
    if path is None:
        return dict(DEFAULT_STRIKE_TABLE)
    table: Dict[int, Tuple[float, float]] = {}
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        # Normalize headers
        if not reader.fieldnames:
            raise SystemExit(f"--strike-table empty or missing headers: {path}")
        field_map = {(_norm_header(h)): h for h in reader.fieldnames if h}
        sku_k = field_map.get("sku")
        target_k = field_map.get("target bph") or field_map.get("target")
        strike_k = field_map.get("strike line") or field_map.get("strike") or field_map.get("strike bph")
        if not sku_k or not strike_k:
            raise SystemExit(
                f"--strike-table needs SKU and Strike Line columns. Found: {reader.fieldnames}"
            )
        for row in reader:
            sku = _parse_sku(row.get(sku_k))
            strike = _num(row.get(strike_k))
            target = _num(row.get(target_k)) if target_k else None
            if sku is None or strike is None:
                continue
            table[sku] = (float(target) if target is not None else float("nan"), float(strike))
    if not table:
        raise SystemExit(f"--strike-table produced no rows: {path}")
    return table


def _norm_header(h: Any) -> str:
    return " ".join(str(h or "").strip().lower().replace("_", " ").split())


def _canon_boxes_headers(headers: Sequence[Any]) -> List[str]:
    aliases = {
        "report date": "Report Date",
        "shift": "Shift",
        "pnp worker name": "Pnp Worker Name",
        "pnp worker": "Pnp Worker Name",
        "worker name": "Pnp Worker Name",
        "station name": "Station Name",
        "station": "Station Name",
        "primary sku": "Primary Sku",
        "boxes packed": "Boxes Packed",
        "packing time seconds": "Packing Time Seconds",
        "packing time (seconds)": "Packing Time Seconds",
    }
    out = []
    for h in headers:
        key = _norm_header(h)
        out.append(aliases.get(key, str(h).strip() if h is not None else ""))
    return out


# ---------------------------------------------------------------------------
# Load Boxes_Packed_by_Worker.xlsx
# ---------------------------------------------------------------------------

def load_boxes_rows(path: Path) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise SystemExit("openpyxl is required to read Boxes_Packed_by_Worker.xlsx (pip install openpyxl)")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not all_rows:
        return []

    # Find header row (first row that looks like Boxes headers)
    header_idx = 0
    for i, row in enumerate(all_rows[:20]):
        keys = {_norm_header(c) for c in row if not _blank(c)}
        if "pnp worker name" in keys and "primary sku" in keys and "boxes packed" in keys:
            header_idx = i
            break

    headers = _canon_boxes_headers(all_rows[header_idx])
    required = ["Pnp Worker Name", "Primary Sku", "Boxes Packed", "Packing Time Seconds", "Shift"]
    missing = [c for c in required if c not in headers]
    if missing:
        raise SystemExit(
            f"{path.name}: missing required column(s): {', '.join(missing)}. "
            f"Found: {[h for h in headers if h]}"
        )

    rows: List[Dict[str, Any]] = []
    for raw in all_rows[header_idx + 1 :]:
        d = {}
        any_val = False
        for i, h in enumerate(headers):
            if not h:
                continue
            v = raw[i] if i < len(raw) else None
            d[h] = v
            if not _blank(v):
                any_val = True
        if not any_val:
            continue  # fully blank pad row

        # Totals row: blank Shift AND blank Pnp Worker Name AND blank Primary Sku
        # (detect — don't hardcode a row number)
        if (
            _blank(d.get("Shift"))
            and _blank(d.get("Pnp Worker Name"))
            and _blank(d.get("Primary Sku"))
        ):
            continue

        rows.append(d)
    return rows


# ---------------------------------------------------------------------------
# Load Raw_Data.csv (pandas)
# ---------------------------------------------------------------------------

@dataclass
class RawSession:
    worker_display: str
    worker_key: str
    first_scan: Any
    last_scan: Any
    idle_pct: Optional[float]
    box_dyn_eff: Optional[float]
    boxes_per_hour: Optional[float]
    total_boxes: Optional[float]
    box_sku_sizes_raw: str
    skus: List[int]
    kind: str  # single_sku | mixed_sku | empty
    session_seconds: Optional[float] = None


def _session_length_seconds(first: Any, last: Any) -> Optional[float]:
    def to_dt(v: Any) -> Optional[datetime]:
        if _blank(v):
            return None
        if isinstance(v, datetime):
            return v
        if pd is not None and isinstance(v, pd.Timestamp):
            return v.to_pydatetime()
        s = str(v).strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%dT%H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
        ):
            try:
                return datetime.strptime(s[:19], fmt) if len(s) >= 16 else datetime.strptime(s, fmt)
            except ValueError:
                continue
        if pd is not None:
            try:
                ts = pd.to_datetime(s, errors="coerce")
                if ts is not None and not pd.isna(ts):
                    return ts.to_pydatetime()
            except Exception:
                return None
        return None

    a = to_dt(first)
    b = to_dt(last)
    if a is None or b is None:
        return None
    return max(0.0, (b - a).total_seconds())


def load_raw_data(path: Path) -> Tuple[List[RawSession], Dict[str, int]]:
    """
    Classify every Raw_Data row as single_sku / mixed_sku.
    Returns (sessions, counts).
    """
    if pd is None:
        raise RuntimeError("pandas is required to read Raw_Data.csv (pip install pandas)")

    # pandas handles CRLF / BOM / stray whitespace in headers better than a hand parser
    df = pd.read_csv(path, encoding="utf-8-sig")
    # Normalize column names
    col_map = {_norm_header(c): c for c in df.columns}
    rename = {}
    want = {
        "pnp worker name": "Pnp Worker Name",
        "first scan": "First Scan",
        "last scan": "Last Scan",
        "idle time %": "Idle Time %",
        "idle time": "Idle Time %",
        "box dynamic efficiency %": "Box Dynamic Efficiency %",
        "boxes per hour": "Boxes per Hour",
        "total boxes packed": "Total Boxes Packed",
        "box sku sizes": "Box Sku Sizes",
        "box sku size": "Box Sku Sizes",
    }
    for key, canon in want.items():
        if key in col_map:
            rename[col_map[key]] = canon
    df = df.rename(columns=rename)

    required = ["Pnp Worker Name", "Box Sku Sizes"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise RuntimeError(
            f"{path.name}: missing column(s): {', '.join(missing)}. "
            f"Found: {list(df.columns)}"
        )

    sessions: List[RawSession] = []
    counts = {"single_sku": 0, "mixed_sku": 0, "empty": 0, "blank_worker": 0}

    for _, row in df.iterrows():
        worker = row.get("Pnp Worker Name")
        if _blank(worker):
            counts["blank_worker"] += 1
            continue
        display = str(worker).strip()
        skus, kind = parse_box_sku_sizes(row.get("Box Sku Sizes"))
        counts[kind] = counts.get(kind, 0) + 1
        first = row.get("First Scan")
        last = row.get("Last Scan")
        sessions.append(
            RawSession(
                worker_display=display,
                worker_key=_norm_name(display),
                first_scan=first,
                last_scan=last,
                idle_pct=_num(row.get("Idle Time %")),
                box_dyn_eff=_num(row.get("Box Dynamic Efficiency %")),
                boxes_per_hour=_num(row.get("Boxes per Hour")),
                total_boxes=_num(row.get("Total Boxes Packed")),
                box_sku_sizes_raw=str(row.get("Box Sku Sizes") or "").strip(),
                skus=skus,
                kind=kind,
                session_seconds=_session_length_seconds(first, last),
            )
        )
    return sessions, counts


# ---------------------------------------------------------------------------
# Step 2 — Primary strike calculation (Boxes only)
# ---------------------------------------------------------------------------

@dataclass
class StrikeRow:
    shift: str
    worker_display: str
    worker_key: str
    station: str
    sku: Optional[int]
    boxes: Optional[float]
    packing_seconds: Optional[float]
    actual_bph: Optional[float]
    target_bph: Optional[float]
    strike_line: Optional[float]
    gap: Optional[float]
    status: str
    reason: str = ""
    single_sku_context: str = ""
    mixed_note: str = ""  # condensed: "yes · 250·600 (3 seg)" or ""


def evaluate_boxes(
    boxes_rows: List[Dict[str, Any]],
    strike_table: Dict[int, Tuple[float, float]],
) -> List[StrikeRow]:
    out: List[StrikeRow] = []
    unknown_warned: set = set()

    for d in boxes_rows:
        shift = "" if _blank(d.get("Shift")) else str(d.get("Shift")).strip()
        worker_raw = d.get("Pnp Worker Name")
        worker_display = "" if _blank(worker_raw) else str(worker_raw).strip()
        worker_key = _norm_name(worker_display) if worker_display else ""
        station = "" if _blank(d.get("Station Name")) else str(d.get("Station Name")).strip()
        sku = _parse_sku(d.get("Primary Sku"))
        boxes = _num(d.get("Boxes Packed"))
        seconds = _num(d.get("Packing Time Seconds"))

        # Bad / incomplete rows — never crash, never silent-drop
        if not worker_display or sku is None:
            reasons = []
            if not worker_display:
                reasons.append("missing worker name")
            if sku is None:
                reasons.append("missing/invalid Primary Sku")
            out.append(
                StrikeRow(
                    shift=shift or "?",
                    worker_display=worker_display or "(blank)",
                    worker_key=worker_key,
                    station=station,
                    sku=sku,
                    boxes=boxes,
                    packing_seconds=seconds,
                    actual_bph=None,
                    target_bph=None,
                    strike_line=None,
                    gap=None,
                    status="bad_row",
                    reason="; ".join(reasons),
                )
            )
            continue

        if boxes is None or seconds is None:
            reasons = []
            if boxes is None:
                reasons.append("missing Boxes Packed")
            if seconds is None:
                reasons.append("missing Packing Time Seconds")
            out.append(
                StrikeRow(
                    shift=shift or "?",
                    worker_display=worker_display,
                    worker_key=worker_key,
                    station=station,
                    sku=sku,
                    boxes=boxes,
                    packing_seconds=seconds,
                    actual_bph=None,
                    target_bph=None,
                    strike_line=None,
                    gap=None,
                    status="bad_row",
                    reason="; ".join(reasons),
                )
            )
            continue

        hours = seconds / 3600.0
        actual_bph = (boxes / hours) if hours > 0 else None

        if sku not in strike_table:
            if sku not in unknown_warned:
                print(
                    f"WARNING: unknown SKU {sku} — no strike line in reference table "
                    f"(will not guess).",
                    file=sys.stderr,
                )
                unknown_warned.add(sku)
            out.append(
                StrikeRow(
                    shift=shift or "?",
                    worker_display=worker_display,
                    worker_key=worker_key,
                    station=station,
                    sku=sku,
                    boxes=boxes,
                    packing_seconds=seconds,
                    actual_bph=actual_bph,
                    target_bph=None,
                    strike_line=None,
                    gap=None,
                    status="unknown_sku",
                    reason=f"SKU {sku} not in strike table",
                )
            )
            continue

        target_bph, strike_line = strike_table[sku]

        if seconds < MIN_PACKING_SECONDS:
            out.append(
                StrikeRow(
                    shift=shift or "?",
                    worker_display=worker_display,
                    worker_key=worker_key,
                    station=station,
                    sku=sku,
                    boxes=boxes,
                    packing_seconds=seconds,
                    actual_bph=actual_bph,
                    target_bph=target_bph,
                    strike_line=strike_line,
                    gap=(actual_bph - strike_line) if actual_bph is not None else None,
                    status="insufficient_data",
                    reason=f"packing time {int(seconds)}s < {MIN_PACKING_SECONDS}s",
                )
            )
            continue

        gap = (actual_bph - strike_line) if actual_bph is not None else None
        if actual_bph is None:
            status = "bad_row"
            reason = "could not compute BPH"
        elif actual_bph < strike_line:
            status = "below_strike_line"
            reason = ""
        else:
            status = "ok"
            reason = ""

        out.append(
            StrikeRow(
                shift=shift or "?",
                worker_display=worker_display,
                worker_key=worker_key,
                station=station,
                sku=sku,
                boxes=boxes,
                packing_seconds=seconds,
                actual_bph=actual_bph,
                target_bph=target_bph,
                strike_line=strike_line,
                gap=gap,
                status=status,
                reason=reason,
            )
        )
    return out


# ---------------------------------------------------------------------------
# Steps 3–4 — attach Raw_Data context (never blend into strike number)
# ---------------------------------------------------------------------------

def _idle_pct_display(idle: Optional[float]) -> Optional[float]:
    """Return idle as 0–100 percent, or None."""
    if idle is None:
        return None
    return idle * 100.0 if idle <= 1.5 else idle


def attach_raw_context(
    strike_rows: List[StrikeRow],
    sessions: List[RawSession],
) -> None:
    """
    Attach condensed Raw_Data context to strike rows.
    Single-SKU → trustworthy SKU context for below_strike candidates.
    Mixed (2+ sizes) → short note only — never used to judge a SKU.
    """
    single = [s for s in sessions if s.kind == "single_sku"]
    mixed = [s for s in sessions if s.kind == "mixed_sku"]

    by_worker_sku: Dict[Tuple[str, int], List[RawSession]] = {}
    for s in single:
        if not s.skus:
            continue
        by_worker_sku.setdefault((s.worker_key, s.skus[0]), []).append(s)

    mixed_by_worker: Dict[str, List[RawSession]] = {}
    for s in mixed:
        mixed_by_worker.setdefault(s.worker_key, []).append(s)

    for row in strike_rows:
        msegs = mixed_by_worker.get(row.worker_key, [])
        if msegs:
            sku_set = set()
            for s in msegs:
                sku_set.update(s.skus)
            sku_txt = "·".join(str(x) for x in sorted(sku_set))
            row.mixed_note = f"yes · {sku_txt} ({len(msegs)} seg)"
        else:
            row.mixed_note = ""

        if row.status != "below_strike_line" or row.sku is None:
            continue

        matches = by_worker_sku.get((row.worker_key, row.sku), [])
        if not matches:
            row.single_sku_context = "—"
            continue
        idles = [x for x in (_idle_pct_display(s.idle_pct) for s in matches) if x is not None]
        bphs = [s.boxes_per_hour for s in matches if s.boxes_per_hour is not None]
        idle_s = f"idle {sum(idles)/len(idles):.0f}%" if idles else "idle ?"
        bph_s = f"BPH {sum(bphs)/len(bphs):.1f}" if bphs else "BPH ?"
        row.single_sku_context = f"{len(matches)}× single · {idle_s} · {bph_s}"


# Back-compat alias used by older tests
def attach_single_sku_context(strike_rows: List[StrikeRow], sessions: List[RawSession]) -> None:
    attach_raw_context(strike_rows, sessions)


def mixed_rollup_by_worker(
    sessions: List[RawSession],
    only_worker_keys: Optional[set] = None,
) -> List[Dict[str, Any]]:
    """
    One row per worker who had mixed-size sessions (2+ SKUs in Box Sku Sizes).
    Condensed — not one row per Raw_Data segment.
    """
    by: Dict[str, Dict[str, Any]] = {}
    for s in sessions:
        if s.kind != "mixed_sku":
            continue
        if only_worker_keys is not None and s.worker_key not in only_worker_keys:
            continue
        if s.worker_key not in by:
            by[s.worker_key] = {
                "worker": s.worker_display,
                "worker_key": s.worker_key,
                "sku_set": set(),
                "segments": 0,
                "boxes": 0.0,
                "idle_vals": [],
                "max_seconds": 0.0,
            }
        w = by[s.worker_key]
        w["segments"] += 1
        w["sku_set"].update(s.skus)
        if s.total_boxes is not None:
            w["boxes"] += s.total_boxes
        idle = _idle_pct_display(s.idle_pct)
        if idle is not None:
            w["idle_vals"].append(idle)
        if s.session_seconds is not None:
            w["max_seconds"] = max(w["max_seconds"], s.session_seconds)

    rows = []
    for w in by.values():
        skus = sorted(w["sku_set"])
        if w["max_seconds"] > 0:
            mins = w["max_seconds"] / 60.0
            length = f"{mins:.0f}m" if mins < 180 else f"{mins/60:.1f}h"
        else:
            length = "?"
        avg_idle = (
            f"{sum(w['idle_vals'])/len(w['idle_vals']):.0f}%"
            if w["idle_vals"]
            else "?"
        )
        rows.append(
            {
                "worker": w["worker"],
                "skus": " · ".join(f"{s}g" for s in skus),
                "sku_count": len(skus),
                "segments": w["segments"],
                "boxes": w["boxes"],
                "idle_pct": avg_idle,
                "span": length,
            }
        )
    rows.sort(key=lambda r: (-r["sku_count"], _norm_name(r["worker"])))
    return rows


def mixed_sessions_table(sessions: List[RawSession]) -> List[Dict[str, Any]]:
    """Alias — condensed rollup (one row per mixed worker)."""
    return mixed_rollup_by_worker(sessions)


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def _fmt(v: Any, digits: int = 1) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:.{digits}f}"
    return str(v)


def print_table(headers: Sequence[str], rows: Sequence[Sequence[Any]], title: str) -> None:
    print()
    print("=" * 78)
    print(title)
    print("=" * 78)
    if not rows:
        print("(none)")
        return
    str_rows = [[_fmt(c) if not isinstance(c, str) else c for c in r] for r in rows]
    widths = [len(h) for h in headers]
    for r in str_rows:
        for i, c in enumerate(r):
            widths[i] = max(widths[i], len(c))
    fmt = "  ".join(f"{{:<{w}}}" for w in widths)
    print(fmt.format(*headers))
    print(fmt.format(*("-" * w for w in widths)))
    for r in str_rows:
        print(fmt.format(*r))


def write_csv_out(path: Path, strike_rows: List[StrikeRow]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "Shift",
                "Worker",
                "SKU",
                "Actual BPH",
                "Strike Line",
                "Gap",
                "Status",
                "Single-SKU context",
                "Mixed sizes (context)",
                "Reason",
            ]
        )
        for r in strike_rows:
            w.writerow(
                [
                    r.shift,
                    r.worker_display,
                    r.sku if r.sku is not None else "",
                    "" if r.actual_bph is None else round(r.actual_bph, 2),
                    "" if r.strike_line is None else r.strike_line,
                    "" if r.gap is None else round(r.gap, 2),
                    r.status,
                    r.single_sku_context,
                    r.mixed_note,
                    r.reason,
                ]
            )


def summarize(
    strike_rows: List[StrikeRow],
    raw_counts: Optional[Dict[str, int]],
    raw_available: bool,
    mixed_workers: int = 0,
) -> None:
    counts = {
        "ok": 0,
        "insufficient_data": 0,
        "below_strike_line": 0,
        "unknown_sku": 0,
        "bad_row": 0,
    }
    for r in strike_rows:
        counts[r.status] = counts.get(r.status, 0) + 1

    print()
    print("=" * 78)
    print("SUMMARY")
    print("=" * 78)
    print(f"  Boxes rows judged:        {len(strike_rows)}")
    print(f"    ok:                     {counts.get('ok', 0)}")
    print(f"    below_strike_line:      {counts.get('below_strike_line', 0)}")
    print(f"    insufficient_data:      {counts.get('insufficient_data', 0)}")
    print(f"    unknown_sku:            {counts.get('unknown_sku', 0)}")
    print(f"    bad_row:                {counts.get('bad_row', 0)}")
    if raw_available and raw_counts is not None:
        print(f"  Raw_Data sessions:")
        print(f"    single_sku:             {raw_counts.get('single_sku', 0)}")
        print(f"    mixed_sku (multi-size): {raw_counts.get('mixed_sku', 0)}")
        print(f"    mixed workers (rolled): {mixed_workers}")
        print(f"    empty Box Sku Sizes:    {raw_counts.get('empty', 0)}")
    else:
        print("  Raw_Data: unavailable this run — mixed/single-SKU context not attached.")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="strike_check",
        description=(
            "Fair strike-candidate report. Boxes Packed = pass/fail. "
            "Raw_Data single-SKU = supporting context; mixed multi-size sessions "
            "rolled up per worker — never blended into a SKU score."
        ),
    )
    p.add_argument(
        "--boxes",
        default="Boxes_Packed_by_Worker.xlsx",
        help="Boxes_Packed_by_Worker.xlsx (default: ./Boxes_Packed_by_Worker.xlsx)",
    )
    p.add_argument(
        "--raw",
        default="Raw_Data.csv",
        help="Raw_Data.csv (default: ./Raw_Data.csv)",
    )
    p.add_argument(
        "--strike-table",
        default=None,
        help="Optional CSV override: SKU, Target BPH, Strike Line",
    )
    p.add_argument(
        "--csv-out",
        default=None,
        help="Write full strike-check table to this CSV path",
    )
    p.add_argument(
        "--all-rows",
        action="store_true",
        help="Print every Boxes row (default: below_strike + insufficient only)",
    )
    p.add_argument(
        "--mixed-all",
        action="store_true",
        help="Show mixed rollup for all workers (default: only workers on the strike list)",
    )
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    boxes_path = Path(args.boxes).expanduser().resolve()
    raw_path = Path(args.raw).expanduser().resolve()
    strike_path = Path(args.strike_table).expanduser().resolve() if args.strike_table else None

    if not boxes_path.is_file():
        print(f"ERROR: Boxes file not found: {boxes_path}", file=sys.stderr)
        return 1

    strike_table = load_strike_table(strike_path)

    try:
        boxes_rows = load_boxes_rows(boxes_path)
    except SystemExit:
        raise
    except Exception as e:
        print(f"ERROR: failed to load Boxes file: {e}", file=sys.stderr)
        return 1

    strike_rows = evaluate_boxes(boxes_rows, strike_table)

    sessions: List[RawSession] = []
    raw_counts: Optional[Dict[str, int]] = None
    raw_available = False
    raw_error: Optional[str] = None

    if not raw_path.is_file():
        raw_error = f"Raw_Data file not found: {raw_path}"
    else:
        try:
            sessions, raw_counts = load_raw_data(raw_path)
            raw_available = True
        except Exception as e:
            raw_error = f"Raw_Data failed to load ({e})"

    if raw_available:
        attach_raw_context(strike_rows, sessions)
    else:
        print(
            f"NOTE: {raw_error}. Printing Boxes strike check only; "
            f"mixed/single-SKU context is unavailable this run.",
            file=sys.stderr,
        )

    focus_statuses = {"below_strike_line", "insufficient_data"}
    if args.all_rows:
        table1 = strike_rows
        title1 = "STRIKE CHECK — all Boxes rows (single SKU per row)"
    else:
        table1 = [r for r in strike_rows if r.status in focus_statuses]
        title1 = "STRIKE CANDIDATES — below strike / insufficient data (Boxes only)"

    print_table(
        [
            "Shift",
            "Worker",
            "SKU",
            "BPH",
            "Strike",
            "Gap",
            "Status",
            "Single context",
            "Mixed sizes?",
        ],
        [
            [
                r.shift.replace("_shift", ""),
                r.worker_display,
                r.sku if r.sku is not None else "",
                round(r.actual_bph, 1) if r.actual_bph is not None else "",
                r.strike_line if r.strike_line is not None else "",
                round(r.gap, 1) if r.gap is not None else "",
                r.status.replace("below_strike_line", "BELOW").replace(
                    "insufficient_data", "short"
                ),
                r.single_sku_context or "—",
                r.mixed_note or "no",
            ]
            for r in table1
        ],
        title1,
    )

    print()
    print("=" * 78)
    print("MIXED SIZES — one row per worker (2+ sizes in Box Sku Sizes)")
    print("Context only — not used to judge any single SKU")
    print("=" * 78)
    if not raw_available:
        print(f"(unavailable — {raw_error})")
        mixed_workers_n = 0
    else:
        if args.mixed_all:
            only_keys = None
            print("(all workers with mixed-size sessions)")
        else:
            only_keys = {r.worker_key for r in table1 if r.worker_key}
            print("(workers on the strike list above only — pass --mixed-all for everyone)")
        mixed_rows = mixed_rollup_by_worker(sessions, only_worker_keys=only_keys)
        mixed_workers_n = len(mixed_rollup_by_worker(sessions))

        if not mixed_rows:
            print("(none for this view)")
        else:
            headers = ["Worker", "Sizes touched", "Segs", "Boxes", "Avg idle", "Span"]
            str_rows = [
                [
                    m["worker"],
                    m["skus"],
                    str(m["segments"]),
                    _fmt(m["boxes"], 0),
                    m["idle_pct"],
                    m["span"],
                ]
                for m in mixed_rows
            ]
            widths = [len(h) for h in headers]
            for r in str_rows:
                for i, c in enumerate(r):
                    widths[i] = max(widths[i], len(str(c)))
            fmt = "  ".join(f"{{:<{w}}}" for w in widths)
            print(fmt.format(*headers))
            print(fmt.format(*("-" * w for w in widths)))
            for r in str_rows:
                print(fmt.format(*[str(c) for c in r]))

    summarize(strike_rows, raw_counts, raw_available, mixed_workers=mixed_workers_n)

    if args.csv_out:
        out = Path(args.csv_out).expanduser().resolve()
        write_csv_out(out, strike_rows)
        print(f"\nWrote full strike-check CSV → {out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
